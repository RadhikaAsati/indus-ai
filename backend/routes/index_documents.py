from fastapi import APIRouter
import os
import csv

from openpyxl import load_workbook

from backend.services.pdf_reader import extract_text
from backend.services.chunker import split_text
from backend.services.embeddings import create_embeddings
from backend.services.vector_store import (
    store_chunks,
    collection
)


router = APIRouter()

UPLOAD_FOLDER = "backend/uploads"

SUPPORTED_EXTENSIONS = (
    ".pdf",
    ".txt",
    ".csv",
    ".xlsx"
)


# ---------------------------------------------------------
# READ TXT FILE
# ---------------------------------------------------------

def read_txt(file_path):
    """
    Read a normal text file.
    """

    with open(
        file_path,
        "r",
        encoding="utf-8"
    ) as file:

        return file.read()


# ---------------------------------------------------------
# READ CSV FILE
# ---------------------------------------------------------

def read_csv(file_path):
    """
    Convert CSV rows into readable structured text.

    Example:

    Machine: HP-05
    Pressure: 210
    Status: Warning
    """

    text_blocks = []

    with open(
        file_path,
        "r",
        encoding="utf-8-sig",
        newline=""
    ) as file:

        reader = csv.DictReader(file)

        for row_number, row in enumerate(
            reader,
            start=1
        ):

            row_lines = [
                f"ROW: {row_number}"
            ]

            for column, value in row.items():

                if value is None:
                    continue

                value = str(value).strip()

                if not value:
                    continue

                row_lines.append(
                    f"{column}: {value}"
                )

            if len(row_lines) > 1:

                text_blocks.append(
                    "\n".join(row_lines)
                )

    return "\n\n".join(
        text_blocks
    )


# ---------------------------------------------------------
# READ EXCEL FILE
# ---------------------------------------------------------

def read_excel(file_path):
    """
    Convert Excel worksheets into readable structured text.

    The first row is treated as the column header.

    Each later row becomes searchable text while preserving
    the worksheet name and column-value relationships.
    """

    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True
    )

    text_blocks = []

    try:

        for sheet in workbook.worksheets:

            rows = sheet.iter_rows(
                values_only=True
            )

            try:
                headers = next(rows)

            except StopIteration:
                continue

            headers = [
                str(header).strip()
                if header is not None
                else f"Column_{index + 1}"

                for index, header in enumerate(
                    headers
                )
            ]

            for row_number, row in enumerate(
                rows,
                start=2
            ):

                row_lines = [
                    f"WORKSHEET: {sheet.title}",
                    f"ROW: {row_number}"
                ]

                has_data = False

                for header, value in zip(
                    headers,
                    row
                ):

                    if value is None:
                        continue

                    value = str(value).strip()

                    if not value:
                        continue

                    row_lines.append(
                        f"{header}: {value}"
                    )

                    has_data = True

                if has_data:

                    text_blocks.append(
                        "\n".join(row_lines)
                    )

    finally:

        workbook.close()

    return "\n\n".join(
        text_blocks
    )


# ---------------------------------------------------------
# READ DOCUMENT BASED ON FORMAT
# ---------------------------------------------------------

def read_document(file_path):
    """
    Select the correct reader based on file type.
    """

    extension = os.path.splitext(
        file_path
    )[1].lower()

    if extension == ".pdf":

        return extract_text(
            file_path
        )

    if extension == ".txt":

        return read_txt(
            file_path
        )

    if extension == ".csv":

        return read_csv(
            file_path
        )

    if extension == ".xlsx":

        return read_excel(
            file_path
        )

    return ""


# ---------------------------------------------------------
# INDEX DOCUMENTS
# ---------------------------------------------------------

@router.post("/index-documents")
def index_documents():
    """
    Index new supported documents into ChromaDB.

    Supported formats:

    - PDF
    - TXT
    - CSV
    - XLSX

    Documents already indexed are skipped automatically.
    """

    indexed_files = []
    skipped_files = []
    failed_files = []

    total_chunks = 0

    # -----------------------------------------------------
    # FIND DOCUMENTS ALREADY IN CHROMA
    # -----------------------------------------------------

    existing_data = collection.get(
        include=["metadatas"]
    )

    existing_documents = set()

    for metadata in existing_data.get(
        "metadatas",
        []
    ):

        document_name = metadata.get(
            "document"
        )

        if document_name:

            existing_documents.add(
                document_name
            )

    # -----------------------------------------------------
    # CHECK FILES IN UPLOAD FOLDER
    # -----------------------------------------------------

    for filename in sorted(
        os.listdir(UPLOAD_FOLDER)
    ):

        extension = os.path.splitext(
            filename
        )[1].lower()

        # Ignore unsupported file types
        if extension not in SUPPORTED_EXTENSIONS:
            continue

        # Skip documents already indexed
        if filename in existing_documents:

            skipped_files.append(
                filename
            )

            continue

        file_path = os.path.join(
            UPLOAD_FOLDER,
            filename
        )

        try:

            # ---------------------------------------------
            # READ
            # ---------------------------------------------

            text = read_document(
                file_path
            )

            if not text.strip():

                failed_files.append(
                    {
                        "file": filename,
                        "reason": "No readable text found."
                    }
                )

                continue

            # ---------------------------------------------
            # CHUNK
            # ---------------------------------------------

            chunks = split_text(
                text
            )

            if not chunks:

                failed_files.append(
                    {
                        "file": filename,
                        "reason": "No chunks were created."
                    }
                )

                continue

            # ---------------------------------------------
            # EMBEDDINGS
            # ---------------------------------------------

            embeddings = create_embeddings(
                chunks
            )

            # ---------------------------------------------
            # STORE IN CHROMA
            # ---------------------------------------------

            store_chunks(
                chunks,
                embeddings,
                filename
            )

            indexed_files.append(
                filename
            )

            total_chunks += len(
                chunks
            )

        except Exception as error:

            # One bad document should not stop
            # the entire indexing process.

            failed_files.append(
                {
                    "file": filename,
                    "reason": str(error)
                }
            )

    # -----------------------------------------------------
    # RESPONSE
    # -----------------------------------------------------

    return {

        "message":
            "Document indexing completed.",

        "supported_formats": [
            "PDF",
            "TXT",
            "CSV",
            "XLSX"
        ],

        "documents_indexed":
            len(indexed_files),

        "documents_skipped":
            len(skipped_files),

        "documents_failed":
            len(failed_files),

        "new_chunks_added":
            total_chunks,

        "indexed_files":
            indexed_files,

        "skipped_files":
            skipped_files,

        "failed_files":
            failed_files
    }
# ---------------------------------------------------------
# LIST INDEXED DOCUMENTS
# ---------------------------------------------------------

@router.get("/documents")
def get_documents():
    """
    Return unique documents currently indexed
    in the INDUS AI Knowledge Brain.
    """

    existing_data = collection.get(
        include=["metadatas"]
    )

    document_chunks = {}

    for metadata in existing_data.get(
        "metadatas",
        []
    ):

        if not metadata:
            continue

        document_name = metadata.get(
            "document"
        )

        if not document_name:
            continue

        if document_name not in document_chunks:

            document_chunks[
                document_name
            ] = 0

        document_chunks[
            document_name
        ] += 1

    documents = []

    for document_name, chunk_count in sorted(
        document_chunks.items()
    ):

        extension = os.path.splitext(
            document_name
        )[1].lower()

        documents.append(
            {
                "filename":
                    document_name,

                "type":
                    extension.replace(
                        ".",
                        ""
                    ).upper()
                    or "DOCUMENT",

                "chunks":
                    chunk_count,

                "status":
                    "indexed"
            }
        )

    return {
        "total_documents":
            len(documents),

        "documents":
            documents
    }